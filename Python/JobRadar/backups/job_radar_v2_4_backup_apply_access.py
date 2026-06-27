import os
import re
import json
import base64
from email.utils import parsedate_to_datetime
from datetime import datetime
from urllib.parse import parse_qs, unquote, urlparse

import pandas as pd
from bs4 import BeautifulSoup

from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build

from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

SCOPES = ["https://www.googleapis.com/auth/gmail.readonly"]

BLACKLIST = [
    "add widget", "manage alerts", "manage job alerts", "view all jobs",
    "unsubscribe", "this email was intended", "linkedin premium",
    "you have an invitation", "thanks for being a valued member",
    "see more jobs", "recommended jobs", "job alert", "your job alert",
    "actively recruiting", "view job", "privacy policy", "email preferences",
]

TARGET_KEYWORDS = {
    "salesforce administrator": 45, "salesforce admin": 45, "salesforce": 35,
    "crm administrator": 35, "crm admin": 35, "crm": 25,
    "business analyst": 35, "analyst": 20, "data analyst": 30,
    "reporting": 25, "dashboard": 20, "analytics": 25,
    "operations analyst": 30, "operations": 20,
    "sales operations": 35, "sales ops": 35,
    "revops": 35, "revenue operations": 35,
    "tableau": 20, "sql": 20, "soql": 20, "excel": 15,
    "d365": 15, "dynamics": 15, "erp": 10,
    "supply chain": 20, "inventory": 20, "materials": 15,
    "procurement": 15, "purchasing": 15, "planner": 15,
    "process improvement": 15, "workflow": 10,
}

MN_CITIES = [
    "minneapolis", "st paul", "saint paul", "bloomington", "hopkins",
    "shoreview", "brooklyn park", "arden hills", "plymouth", "minnetonka",
    "roseville", "eden prairie", "maple grove", "st. louis park", "wayzata",
]

INTERNATIONAL_LOCATIONS = [
    "india", "australia", "france", "japan", "nigeria", "mexico",
    "netherlands", "indonesia", "united kingdom", "south korea",
    "korea", "guam", "philippines", "canada", "ireland", "germany",
]

TARGET_COMPANIES = {
    "optum": 25,
    "unitedhealth": 25,
    "unitedhealth group": 25,
    "cargill": 20,
    "best buy": 20,
    "thermo fisher": 20,
    "salesforce": 25,
    "medica": 20,
    "wells fargo": 20,
    "patterson": 20,
    "patterson companies": 20,
    "3m": 15,
    "target": 15,
    "mayo clinic": 15,
    "donaldson": 15,
    "boston scientific": 15,
    "jack link": 15,
}

TRACKER_FILE = "data/application_tracker.csv"
SETTINGS_FILE = "data/settings.json"
COMPANY_WATCHLIST_FILE = "data/company_watchlist.csv"
RECRUITER_CRM_FILE = "data/recruiter_crm.csv"

TRACKER_COLUMNS = [
    "job_key",
    "job_title",
    "company",
    "status",
    "application_stage",
    "date_applied",
    "follow_up",
    "last_updated",
    "resume_used",
    "cover_letter_used",
    "recruiter_name",
    "recruiter_email",
    "interview_date",
    "salary",
    "job_link",
    "apply_url",
    "source_email_url",
    "resume_file_suggestion",
    "resume_test_group",
    "notes",
]

RECRUITER_CRM_COLUMNS = [
    "recruiter_key",
    "recruiter_name",
    "company",
    "email",
    "lead_type",
    "priority",
    "status",
    "last_contact_date",
    "next_follow_up",
    "relationship_score",
    "jobs_shared",
    "latest_subject",
    "suggested_action",
    "notes",
]

VALID_APPLICATION_STATUSES = [
    "Not Applied",
    "Applied",
    "Recruiter Contacted",
    "Phone Screen",
    "Interview Scheduled",
    "Interview Complete",
    "Final Interview",
    "Offer",
    "Accepted",
    "Rejected",
    "Withdrawn",
]

DEFAULT_SETTINGS = {
    "preferred_locations": ["remote", "hybrid", "minnesota", "united states"],
    "preferred_states": ["MN", "Minnesota"],
    "avoid_locations": INTERNATIONAL_LOCATIONS,
    "minimum_salary": 80000,
    "prefer_remote": True,
    "prefer_hybrid": True,
    "allow_onsite": False,
    "preferred_roles": [
        "Salesforce Administrator",
        "CRM Administrator",
        "Business Analyst",
        "Sales Operations Analyst",
        "RevOps Analyst",
        "Operations Analyst",
        "Data Analyst"
    ]
}

DEFAULT_COMPANY_WATCHLIST = [
    {"company": "Salesforce", "bonus": 35, "category": "Dream Company", "notes": "Salesforce ecosystem target"},
    {"company": "Optum", "bonus": 25, "category": "Target Company", "notes": "Healthcare and operations fit"},
    {"company": "UnitedHealth Group", "bonus": 25, "category": "Target Company", "notes": "Healthcare operations fit"},
    {"company": "Medica", "bonus": 20, "category": "Target Company", "notes": "Health insurance background fit"},
    {"company": "Best Buy", "bonus": 20, "category": "Target Company", "notes": "Minnesota employer"},
    {"company": "Cargill", "bonus": 20, "category": "Target Company", "notes": "Minnesota employer"},
    {"company": "Thermo Fisher", "bonus": 20, "category": "Target Company", "notes": "Operations and analytics fit"},
    {"company": "Patterson Companies", "bonus": 20, "category": "Target Company", "notes": "Salesforce/admin fit"},
    {"company": "Target", "bonus": 15, "category": "Good Company", "notes": "Minnesota employer"},
    {"company": "3M", "bonus": 15, "category": "Good Company", "notes": "Minnesota employer"},
]


def load_settings():
    """Load user-editable settings. Create default settings if missing."""
    os.makedirs("data", exist_ok=True)

    if not os.path.exists(SETTINGS_FILE):
        with open(SETTINGS_FILE, "w", encoding="utf-8") as file:
            json.dump(DEFAULT_SETTINGS, file, indent=4)
        return DEFAULT_SETTINGS

    with open(SETTINGS_FILE, "r", encoding="utf-8") as file:
        loaded_settings = json.load(file)

    settings = DEFAULT_SETTINGS.copy()
    settings.update(loaded_settings)
    return settings


def load_company_watchlist():
    """Load user-editable company scoring list. Create default watchlist if missing."""
    os.makedirs("data", exist_ok=True)

    if not os.path.exists(COMPANY_WATCHLIST_FILE):
        watchlist_df = pd.DataFrame(DEFAULT_COMPANY_WATCHLIST)
        watchlist_df.to_csv(COMPANY_WATCHLIST_FILE, index=False)
        return watchlist_df

    watchlist_df = pd.read_csv(COMPANY_WATCHLIST_FILE, dtype=str).fillna("")

    for column in ["company", "bonus", "category", "notes"]:
        if column not in watchlist_df.columns:
            watchlist_df[column] = ""

    return watchlist_df[["company", "bonus", "category", "notes"]]


def connect_gmail():
    creds = None

    if os.path.exists("token.json"):
        creds = Credentials.from_authorized_user_file("token.json", SCOPES)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file("credentials.json", SCOPES)
            creds = flow.run_local_server(port=0)

        with open("token.json", "w") as token:
            token.write(creds.to_json())

    return build("gmail", "v1", credentials=creds)


def get_header(headers, name):
    for header in headers:
        if header["name"].lower() == name.lower():
            return header["value"]
    return ""


def search_job_alert_emails(service):
    query = (
        "("
        "from:jobalerts-noreply@linkedin.com OR "
        "from:jobalerts@sites.careerbuilder.com OR "
        "from:notify-noreply@google.com OR "
        "from:googlealerts-noreply@google.com OR "
        "from:alerts@indeed.com OR "
        "from:noreply@indeed.com OR "
        "from:ziprecruiter.com"
        ") newer_than:14d"
    )

    results = service.users().messages().list(
        userId="me",
        q=query,
        maxResults=100
    ).execute()

    return results.get("messages", [])


def search_recruiter_emails(service):
    query = (
        "("
        "from:messages-noreply@linkedin.com OR "
        "from:messaging-digest-noreply@linkedin.com OR "
        "from:invitations@linkedin.com OR "
        "from:linkedin.com"
        ") "
        "("
        "recruiter OR hiring OR opportunity OR interview OR "
        '"messaged you" OR "sent you an invitation" OR "waiting for your response"'
        ") newer_than:14d"
    )

    results = service.users().messages().list(
        userId="me",
        q=query,
        maxResults=50
    ).execute()

    return results.get("messages", [])


def decode_body(data):
    if not data:
        return ""

    data = data.replace("-", "+").replace("_", "/")
    decoded_bytes = base64.b64decode(data)
    return decoded_bytes.decode("utf-8", errors="ignore")


def get_email_body(payload):
    body_text = ""

    if "parts" in payload:
        for part in payload["parts"]:
            mime_type = part.get("mimeType", "")
            body_data = part.get("body", {}).get("data")

            if mime_type in ["text/html", "text/plain"] and body_data:
                body_text += decode_body(body_data)

            if "parts" in part:
                body_text += get_email_body(part)
    else:
        body_data = payload.get("body", {}).get("data")
        if body_data:
            body_text += decode_body(body_data)

    return body_text



def clean_tracking_url(url):
    """Unwrap common redirect/tracking URLs so the Apply link is easier to use."""
    url = str(url or "").strip()
    if not url:
        return ""

    if url.startswith("//"):
        url = "https:" + url

    parsed = urlparse(url)
    query = parse_qs(parsed.query)

    # Google and newsletter redirect links often store the real destination here.
    for key in ["url", "u", "q", "target", "redirect", "redirect_url"]:
        if key in query and query[key]:
            candidate = unquote(query[key][0])
            if candidate.startswith("http"):
                return candidate

    return url


def is_useful_apply_url(url):
    """Keep links that are likely to lead to a job/apply page and drop unsubscribe/settings links."""
    url_lower = str(url or "").lower()
    if not url_lower.startswith("http"):
        return False

    bad_terms = [
        "unsubscribe", "email-preferences", "preferences", "privacy", "terms",
        "help", "support", "learning", "premium", "utm_medium=email_preferences",
        "managealerts", "manage-alerts", "salary", "share"
    ]
    if any(term in url_lower for term in bad_terms):
        return False

    good_terms = [
        "linkedin.com/jobs", "indeed.com", "ziprecruiter", "careerbuilder",
        "greenhouse.io", "lever.co", "myworkdayjobs", "workdayjobs",
        "smartrecruiters", "icims", "jobvite", "ashbyhq", "bamboohr",
        "careers", "jobs", "apply", "jobdetails", "viewjob", "job-detail"
    ]
    return any(term in url_lower for term in good_terms)


def extract_links_from_html(raw_body):
    """Extract and clean useful links from the original HTML email body."""
    soup = BeautifulSoup(raw_body or "", "html.parser")
    links = []
    seen = set()

    for anchor in soup.find_all("a"):
        href = anchor.get("href", "")
        link_text = clean_export_text(anchor.get_text(" "))
        clean_href = clean_tracking_url(href)

        if not is_useful_apply_url(clean_href):
            continue

        if clean_href in seen:
            continue

        seen.add(clean_href)
        links.append({
            "url": clean_href,
            "text": link_text,
        })

    return links


def choose_apply_url(email, job_title="", company=""):
    """Pick the best apply URL found in the email for this job."""
    links = email.get("links", []) or []
    if not links:
        return ""

    title_tokens = [token for token in re.findall(r"[a-z0-9]+", str(job_title).lower()) if len(token) >= 4]
    company_tokens = [token for token in re.findall(r"[a-z0-9]+", str(company).lower()) if len(token) >= 4]

    scored_links = []
    for link in links:
        url = link.get("url", "")
        link_text = link.get("text", "")
        combined = f"{url} {link_text}".lower()
        score = 0

        if "linkedin.com/jobs" in combined:
            score += 50
        if any(term in combined for term in ["apply", "view job", "job", "career", "details"]):
            score += 25
        for token in title_tokens[:6]:
            if token in combined:
                score += 5
        for token in company_tokens[:4]:
            if token in combined:
                score += 8

        scored_links.append((score, url))

    scored_links.sort(reverse=True, key=lambda item: item[0])
    return scored_links[0][1] if scored_links else ""


def build_gmail_url(message_id):
    if not message_id:
        return ""
    return f"https://mail.google.com/mail/u/0/#all/{message_id}"

def read_full_email(service, message_id):
    message = service.users().messages().get(
        userId="me",
        id=message_id,
        format="full"
    ).execute()

    headers = message["payload"]["headers"]

    subject = get_header(headers, "Subject")
    sender = get_header(headers, "From")
    date_raw = get_header(headers, "Date")

    try:
        email_date = parsedate_to_datetime(date_raw).strftime("%Y-%m-%d")
    except Exception:
        email_date = date_raw

    raw_body = get_email_body(message["payload"])
    links = extract_links_from_html(raw_body)
    soup = BeautifulSoup(raw_body, "html.parser")
    clean_text = soup.get_text("\n")

    return {
        "subject": subject,
        "sender": sender,
        "date": email_date,
        "body": clean_text,
        "snippet": message.get("snippet", ""),
        "message_id": message_id,
        "email_url": build_gmail_url(message_id),
        "links": links,
    }


def is_junk_title(title):
    title_lower = title.lower().strip()

    if len(title_lower) < 4:
        return True

    if set(title_lower) <= {"-", "_", "—", " "}:
        return True

    if title_lower.count("-") > 10:
        return True

    return any(bad_word in title_lower for bad_word in BLACKLIST)


def build_job(job_title, company, location, source, email):
    apply_url = choose_apply_url(email, job_title, company)
    return {
        "job_title": job_title.strip(),
        "company": company.strip(),
        "location": location.strip(),
        "source": source,
        "apply_url": apply_url,
        "job_link": apply_url,
        "source_email_url": email.get("email_url", ""),
        "apply_link_status": "Apply Link Found" if apply_url else "Open Source Email",
        "email_subject": email["subject"],
        "email_date": email["date"]
    }


def extract_linkedin_jobs(email):
    jobs = []
    lines = [line.strip() for line in email["body"].splitlines() if line.strip()]

    for i, line in enumerate(lines):
        if is_junk_title(line):
            continue

        next_line = lines[i + 1] if i + 1 < len(lines) else ""

        if "·" in next_line:
            parts = [p.strip() for p in next_line.split("·")]
            company = parts[0] if len(parts) > 0 else ""
            location = parts[1] if len(parts) > 1 else ""

            if not is_junk_title(company):
                jobs.append(build_job(line, company, location, "LinkedIn", email))

    return jobs


def extract_google_jobs(email):
    jobs = []
    lines = [line.strip() for line in email["body"].splitlines() if line.strip()]

    for i, line in enumerate(lines):
        if is_junk_title(line):
            continue

        if any(keyword in line.lower() for keyword in [
            "analyst", "administrator", "salesforce", "crm", "operations",
            "data", "reporting", "manager", "coordinator", "specialist",
            "inventory", "supply chain"
        ]):
            company = lines[i + 1] if i + 1 < len(lines) else ""
            location = lines[i + 2] if i + 2 < len(lines) else ""

            if not is_junk_title(company):
                jobs.append(build_job(line, company, location, "Google Job Alert", email))

    return jobs


def extract_simple_jobs(email):
    jobs = []
    lines = [line.strip() for line in email["body"].splitlines() if line.strip()]

    for line in lines:
        if is_junk_title(line):
            continue

        if any(keyword in line.lower() for keyword in TARGET_KEYWORDS.keys()):
            jobs.append(build_job(line, "", "", "Other", email))

    return jobs


def extract_recruiter_lead(email):
    subject = email.get("subject", "")
    sender = email.get("sender", "")
    snippet = email.get("snippet", "")
    body = email.get("body", "")

    combined = " ".join([subject, sender, snippet, body]).lower()

    if "job alert" in combined or "recommended jobs" in combined:
        return None

    if not any(word in combined for word in [
        "recruiter", "hiring", "opportunity", "interview",
        "messaged you", "sent you an invitation", "waiting for your response"
    ]):
        return None

    lead_type = "LinkedIn Message"
    if "invitation" in combined or "waiting for your response" in combined:
        lead_type = "LinkedIn Invitation"

    priority = "Review"
    if any(word in combined for word in ["recruiter", "hiring", "opportunity", "interview"]):
        priority = "High Priority"

    return {
        "lead_type": lead_type,
        "sender": clean_export_text(sender),
        "subject": clean_export_text(subject),
        "date": email.get("date", ""),
        "priority": priority,
        "status": "Not Reviewed",
        "snippet": clean_export_text(snippet)
    }




def extract_email_address(value):
    """Pull an email address from a sender string when available."""
    value = str(value or "")
    match = re.search(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}", value)
    return match.group(0).lower() if match else ""


def guess_recruiter_name(sender):
    """Create a readable recruiter name from the Gmail sender field."""
    sender = clean_export_text(sender)
    sender = re.sub(r"<.*?>", "", sender).strip()
    sender = sender.replace('"', '').strip()

    if "@" in sender and not re.search(r"\s", sender):
        return sender.split("@")[0].replace(".", " ").replace("_", " ").title()

    return sender or "Unknown Recruiter"


def guess_recruiter_company(sender, subject="", snippet=""):
    """Infer company from sender/domain or recognizable text."""
    combined = " ".join([str(sender or ""), str(subject or ""), str(snippet or "")]).lower()

    known_companies = [
        "linkedin", "motion recruitment", "robert half", "dice", "indeed",
        "ziprecruiter", "careerbuilder", "optum", "unitedhealth", "salesforce",
        "best buy", "cargill", "thermo fisher", "medica", "patterson"
    ]

    for company in known_companies:
        if company in combined:
            return company.title()

    email = extract_email_address(sender)
    if email and "@" in email:
        domain = email.split("@")[1].split(".")[0]
        return domain.replace("-", " ").title()

    return "Unknown"


def create_recruiter_key(lead):
    """Create a stable key for recruiter CRM tracking."""
    email = extract_email_address(lead.get("sender", ""))
    if email:
        return email

    sender = clean_export_text(lead.get("sender", "")).lower()
    return re.sub(r"[^a-z0-9]+", "_", sender).strip("_") or "unknown_recruiter"


def load_recruiter_crm():
    """Load recruiter CRM or create a blank CRM file."""
    os.makedirs("data", exist_ok=True)

    if not os.path.exists(RECRUITER_CRM_FILE):
        crm_df = pd.DataFrame(columns=RECRUITER_CRM_COLUMNS)
        crm_df.to_csv(RECRUITER_CRM_FILE, index=False)
        return crm_df

    crm_df = pd.read_csv(RECRUITER_CRM_FILE, dtype=str).fillna("")

    for column in RECRUITER_CRM_COLUMNS:
        if column not in crm_df.columns:
            crm_df[column] = ""

    crm_df = crm_df[RECRUITER_CRM_COLUMNS]
    crm_df = crm_df.drop_duplicates(subset=["recruiter_key"], keep="last")
    return crm_df


def save_recruiter_crm(crm_df):
    """Save recruiter CRM data."""
    os.makedirs("data", exist_ok=True)
    crm_df = crm_df[RECRUITER_CRM_COLUMNS]
    crm_df.to_csv(RECRUITER_CRM_FILE, index=False)


def get_recruiter_suggested_action(priority, status):
    priority = str(priority or "").lower()
    status = str(status or "").lower()

    if status in ["new", "not reviewed"] and "high" in priority:
        return "Review and respond today"
    if status in ["waiting", "contacted"]:
        return "Follow up if no response within 2 business days"
    if status in ["responded", "active"]:
        return "Keep warm and track next step"
    return "Review when time allows"


def apply_recruiter_crm(recruiter_df):
    """Merge recruiter leads into a persistent recruiter CRM."""
    crm_df = load_recruiter_crm()
    today = datetime.now().strftime("%Y-%m-%d")

    if recruiter_df.empty:
        return recruiter_df, crm_df

    crm_lookup = {
        row["recruiter_key"]: row.to_dict()
        for _, row in crm_df.iterrows()
        if row.get("recruiter_key", "")
    }

    new_rows = []
    enriched_rows = []

    for _, lead in recruiter_df.iterrows():
        lead_dict = lead.to_dict()
        recruiter_key = create_recruiter_key(lead_dict)
        saved = crm_lookup.get(recruiter_key)

        recruiter_name = guess_recruiter_name(lead_dict.get("sender", ""))
        company = guess_recruiter_company(
            lead_dict.get("sender", ""),
            lead_dict.get("subject", ""),
            lead_dict.get("snippet", "")
        )
        email = extract_email_address(lead_dict.get("sender", ""))
        priority = lead_dict.get("priority", "Review")
        lead_type = lead_dict.get("lead_type", "Recruiter Lead")

        if saved:
            status = saved.get("status", "New") or "New"
            relationship_score = saved.get("relationship_score", "") or "50"
            jobs_shared = str(int(saved.get("jobs_shared", "0") or 0) + 1)
            notes = saved.get("notes", "")
            next_follow_up = saved.get("next_follow_up", "")
        else:
            status = "New"
            relationship_score = "75" if priority == "High Priority" else "50"
            jobs_shared = "1"
            notes = ""
            next_follow_up = ""

            new_rows.append({
                "recruiter_key": recruiter_key,
                "recruiter_name": recruiter_name,
                "company": company,
                "email": email,
                "lead_type": lead_type,
                "priority": priority,
                "status": status,
                "last_contact_date": lead_dict.get("date", today),
                "next_follow_up": next_follow_up,
                "relationship_score": relationship_score,
                "jobs_shared": jobs_shared,
                "latest_subject": lead_dict.get("subject", ""),
                "suggested_action": get_recruiter_suggested_action(priority, status),
                "notes": notes,
            })

        lead_dict["recruiter_key"] = recruiter_key
        lead_dict["recruiter_name"] = recruiter_name
        lead_dict["company"] = company
        lead_dict["email"] = email
        lead_dict["crm_status"] = status
        lead_dict["relationship_score"] = relationship_score
        lead_dict["suggested_action"] = get_recruiter_suggested_action(priority, status)
        enriched_rows.append(lead_dict)

    if new_rows:
        crm_df = pd.concat([crm_df, pd.DataFrame(new_rows)], ignore_index=True)
        crm_df = crm_df.drop_duplicates(subset=["recruiter_key"], keep="last")
        save_recruiter_crm(crm_df)

    return pd.DataFrame(enriched_rows), load_recruiter_crm()

def has_salary_signal(text):
    return bool(
        re.search(r"\$\s?(80|90|100|110|120|130|140|150|160|170|180|190|200)\s?k", text)
        or re.search(r"\$\s?(80,000|90,000|100,000|110,000|120,000|130,000|140,000|150,000)", text)
        or re.search(r"up to \$\s?\d+", text)
    )


def get_target_track(job):
    text = " ".join([
        job.get("job_title", ""),
        job.get("company", ""),
        job.get("location", ""),
        job.get("email_subject", ""),
        job.get("apply_url", "")
    ]).lower()

    if any(word in text for word in ["salesforce", "crm"]):
        return "Salesforce / CRM"

    if any(word in text for word in ["sales operations", "sales ops", "revops", "revenue operations"]):
        return "Sales Ops / RevOps"

    if any(word in text for word in ["business analyst", "project coordinator", "process analyst", "product owner"]):
        return "Business Analyst"

    if any(word in text for word in ["data analyst", "analytics", "reporting", "dashboard", "tableau", "sql", "excel"]):
        return "Data / Reporting"

    if any(word in text for word in [
        "operations", "supply chain", "inventory", "materials",
        "procurement", "purchasing", "planner", "logistics"
    ]):
        return "Operations / Supply Chain"

    return "General / Other"


def get_recommended_resume(target_track):
    if target_track == "Salesforce / CRM":
        return "Salesforce Resume"
    if target_track == "Sales Ops / RevOps":
        return "Sales Ops / RevOps Resume"
    if target_track == "Business Analyst":
        return "Business Analyst Resume"
    if target_track == "Data / Reporting":
        return "Business Analytics Resume"
    if target_track == "Operations / Supply Chain":
        return "Operations / Supply Chain Resume"
    return "General Resume"


def get_resume_advisor(job):
    """Create rule-based resume advice without calling an AI API yet."""
    text = " ".join([
        job.get("job_title", ""),
        job.get("company", ""),
        job.get("location", ""),
        job.get("email_subject", ""),
        job.get("match_reasons", "")
    ]).lower()

    target_track = job.get("target_track", "General / Other")
    score = int(job.get("score", 0) or 0)

    suggestions = []
    missing_skills = []
    interview_focus = []

    if target_track == "Salesforce / CRM":
        suggestions.extend([
            "Lead with Salesforce Administrator certification",
            "Highlight Flow Builder, reports, dashboards, and CRM process improvement",
            "Mention hands-on Salesforce project work and business problem solving"
        ])
        interview_focus.extend([
            "Salesforce security model",
            "Flow Builder examples",
            "Reports and dashboards",
            "Business requirements gathering"
        ])
        for skill in ["apex", "data cloud", "cpq", "agentforce", "service cloud"]:
            if skill in text:
                missing_skills.append(skill.title())

    elif target_track == "Business Analyst":
        suggestions.extend([
            "Use the Business Analyst resume",
            "Highlight requirements gathering, stakeholder communication, and dashboard work",
            "Connect analytics projects to business decisions"
        ])
        interview_focus.extend([
            "Requirements gathering",
            "Stakeholder communication",
            "Data storytelling",
            "Process improvement"
        ])
        for skill in ["jira", "agile", "scrum", "user stories", "uat"]:
            if skill in text:
                missing_skills.append(skill.title())

    elif target_track == "Sales Ops / RevOps":
        suggestions.extend([
            "Use the Sales Ops / RevOps resume",
            "Highlight sales operations, CRM, reporting, and process improvement",
            "Connect sales background to operational reporting and pipeline support"
        ])
        interview_focus.extend([
            "Pipeline reporting",
            "Sales process improvement",
            "CRM data quality",
            "Stakeholder support"
        ])
        for skill in ["forecasting", "hubspot", "salesloft", "outreach", "revenue operations"]:
            if skill in text:
                missing_skills.append(skill.title())

    elif target_track == "Operations / Supply Chain":
        suggestions.extend([
            "Use the Operations / Supply Chain resume",
            "Highlight inventory, purchasing, work orders, Maximo, and plant operations",
            "Connect storeroom experience to process improvement and data tracking"
        ])
        interview_focus.extend([
            "Inventory control",
            "Purchasing and receiving",
            "Process improvement",
            "Operations problem solving"
        ])
        for skill in ["sap", "oracle", "mrp", "erp", "forecasting"]:
            if skill in text:
                missing_skills.append(skill.upper() if skill in ["sap", "mrp", "erp"] else skill.title())

    elif target_track == "Data / Reporting":
        suggestions.extend([
            "Use the Business Analytics resume",
            "Highlight Excel, dashboards, reporting, Tableau, and analytics coursework",
            "Show how reports helped decision making"
        ])
        interview_focus.extend([
            "Excel analysis",
            "Dashboard design",
            "Data storytelling",
            "SQL/SOQL basics"
        ])
        for skill in ["python", "power bi", "sql", "tableau", "looker"]:
            if skill in text:
                missing_skills.append(skill.upper() if skill == "sql" else skill.title())

    else:
        suggestions.extend([
            "Use the General resume",
            "Highlight transferable operations, sales, customer support, and problem solving experience"
        ])
        interview_focus.extend([
            "Transferable skills",
            "Customer and stakeholder communication",
            "Problem solving"
        ])

    if score >= 120:
        resume_fit_level = "Excellent Resume Fit"
    elif score >= 90:
        resume_fit_level = "Strong Resume Fit"
    elif score >= 60:
        resume_fit_level = "Possible Resume Fit"
    else:
        resume_fit_level = "Low Resume Fit"

    # Keep the spreadsheet readable.
    return {
        "resume_fit_level": resume_fit_level,
        "resume_fit_score": min(100, score),
        "resume_suggestions": " | ".join(dict.fromkeys(suggestions)),
        "missing_or_watch_skills": " | ".join(dict.fromkeys(missing_skills)) if missing_skills else "None flagged",
        "interview_focus": " | ".join(dict.fromkeys(interview_focus)),
    }



def get_job_intelligence(job):
    """Create rule-based job intelligence without using a paid AI API."""
    title = job.get("job_title", "")
    company = job.get("company", "") or "this company"
    target_track = job.get("target_track", "General / Other")
    recommended_resume = job.get("recommended_resume", "General Resume")
    location_status = job.get("location_status", "Unknown")
    company_category = job.get("company_watchlist_category", "")
    score = int(job.get("score", 0) or 0)
    fit = job.get("fit_category", "")
    resume_fit = job.get("resume_fit_level", "")
    reasons = job.get("match_reasons", "")
    missing = job.get("missing_or_watch_skills", "None flagged")

    fit_points = []
    concerns = []

    if "salesforce" in target_track.lower() or "crm" in target_track.lower():
        fit_points.append("Aligns with Salesforce Administrator and CRM career goals")
        fit_points.append("Good opportunity to discuss Salesforce projects, dashboards, and automation")
    elif "business analyst" in target_track.lower():
        fit_points.append("Aligns with business analysis, requirements gathering, and data storytelling goals")
        fit_points.append("Good opportunity to connect projects to business decision-making")
    elif "sales ops" in target_track.lower() or "revops" in target_track.lower():
        fit_points.append("Aligns with sales operations, CRM reporting, and pipeline support experience")
        fit_points.append("Strong bridge between sales background and technical CRM skills")
    elif "operations" in target_track.lower() or "supply chain" in target_track.lower():
        fit_points.append("Aligns with operations, inventory, purchasing, and process improvement experience")
        fit_points.append("Good fit for connecting plant operations experience to analytics and reporting")
    elif "data" in target_track.lower() or "reporting" in target_track.lower():
        fit_points.append("Aligns with reporting, dashboards, Excel, Tableau, and analytics development")
        fit_points.append("Good opportunity to discuss data storytelling and decision support")
    else:
        fit_points.append("Possible transferable-skills opportunity")

    if location_status == "Preferred":
        fit_points.append("Matches preferred location/work arrangement")
    elif location_status == "Outside Target Area":
        concerns.append("Location appears outside the target area")
    elif location_status == "On-site":
        concerns.append("Role appears on-site, which may not match hybrid/remote preference")
    else:
        concerns.append("Location may need manual review")

    if company_category:
        fit_points.append(f"Company is flagged as {company_category}")

    if missing and missing != "None flagged":
        concerns.append(f"Review possible missing/watch skills: {missing}")

    if score >= 90:
        next_action = "Prioritize for application"
    elif score >= 70:
        next_action = "Review soon and compare against current application goals"
    elif score >= 50:
        next_action = "Review only if time allows"
    else:
        next_action = "Do not prioritize unless there is a strategic reason"

    if not concerns:
        concerns.append("No major concerns flagged by current rules")

    why_this_fits = f"{title} at {company} is a {fit.lower()} for the {target_track} track. " + " ".join(fit_points)
    resume_focus = f"Use {recommended_resume}. Focus on: {job.get('resume_suggestions', '')}"
    interview_focus = job.get("interview_focus", "Review role requirements and prepare project examples")

    return {
        "job_intelligence_summary": clean_export_text(why_this_fits),
        "why_this_fits": clean_export_text(" | ".join(fit_points)),
        "resume_focus": clean_export_text(resume_focus),
        "job_intelligence_interview_focus": clean_export_text(interview_focus),
        "potential_concerns": clean_export_text(" | ".join(concerns)),
        "suggested_next_action": next_action,
    }

def get_match_percent(score):
    return min(100, score)


def get_fit_category(score):
    if score >= 90:
        return "Excellent Match"
    if score >= 70:
        return "Strong Match"
    if score >= 50:
        return "Possible Match"
    return "Low Match"


def get_recommendation(score):
    if score >= 90:
        return "Apply Now"
    if score >= 70:
        return "High Priority"
    if score >= 50:
        return "Review"
    return "Skip"


def get_apply_priority(score):
    if score >= 90:
        return "Apply within 24 hours"
    if score >= 70:
        return "Apply this week"
    if score >= 50:
        return "Review if time"
    return "Do not prioritize"


def clean_export_text(value):
    if value is None:
        return ""

    value = str(value)
    value = value.replace("â€™", "'")
    value = value.replace("â€œ", '"')
    value = value.replace("â€", '"')
    value = value.replace("â€“", "-")
    value = value.replace("Â", "")
    value = re.sub(r"[À-ÿ]{5,}", "", value)
    value = re.sub(r"\s+", " ", value).strip()
    return value


def get_location_status(job_text):
    settings = load_settings()
    text = job_text.lower()

    avoid_locations = [location.lower() for location in settings.get("avoid_locations", INTERNATIONAL_LOCATIONS)]
    preferred_locations = [location.lower() for location in settings.get("preferred_locations", [])]

    if any(location in text for location in avoid_locations):
        return "Outside Target Area"

    if "remote" in text or "work from home" in text:
        return "Preferred" if settings.get("prefer_remote", True) else "Review"

    if "hybrid" in text:
        return "Preferred" if settings.get("prefer_hybrid", True) else "Review"

    if any(location in text for location in preferred_locations):
        return "Preferred"

    if any(city in text for city in MN_CITIES) or ", mn" in text or "minnesota" in text:
        return "Preferred"

    if "united states" in text or "usa" in text:
        return "Preferred"

    if "on-site" in text or "onsite" in text or "on site" in text:
        return "On-site" if not settings.get("allow_onsite", False) else "Review"

    return "Unknown"


def get_company_bonus(company_text):
    text = company_text.lower()
    watchlist_df = load_company_watchlist()

    for _, row in watchlist_df.iterrows():
        company = str(row.get("company", "")).lower().strip()
        if company and company in text:
            try:
                points = int(float(row.get("bonus", 0)))
            except Exception:
                points = 0

            category = row.get("category", "Watchlist Company") or "Watchlist Company"
            return points, row.get("company", ""), category

    for company, points in TARGET_COMPANIES.items():
        if company in text:
            return points, company, "Target Company"

    return 0, "", ""


def get_target_strength(job):
    text = " ".join([
        job.get("job_title", ""),
        job.get("company", ""),
        job.get("location", ""),
        job.get("email_subject", "")
    ]).lower()

    strong_terms = [
        "salesforce administrator", "salesforce admin", "crm administrator",
        "business analyst", "sales operations analyst", "revops",
        "revenue operations", "operations analyst", "data analyst"
    ]

    good_terms = [
        "salesforce", "crm", "sales operations", "analytics",
        "reporting", "dashboard", "operations", "supply chain", "inventory"
    ]

    if any(term in text for term in strong_terms):
        return "Strong Target"

    if any(term in text for term in good_terms):
        return "Good Target"

    return "Standard"



def get_resume_file_suggestion(target_track):
    """Suggest a practical resume file/version to test for callback tracking."""
    if target_track == "Salesforce / CRM":
        return "Domonique_Robinson_Salesforce_CRM_Resume"
    if target_track == "Sales Ops / RevOps":
        return "Domonique_Robinson_SalesOps_RevOps_Resume"
    if target_track == "Business Analyst":
        return "Domonique_Robinson_Business_Analyst_Resume"
    if target_track == "Data / Reporting":
        return "Domonique_Robinson_Business_Analytics_Resume"
    if target_track == "Operations / Supply Chain":
        return "Domonique_Robinson_Operations_SupplyChain_Resume"
    return "Domonique_Robinson_Master_Resume"


def get_resume_test_group(target_track):
    """Make it easier to compare which resume version gets callbacks."""
    if target_track == "Salesforce / CRM":
        return "Resume Test A - Salesforce CRM"
    if target_track == "Sales Ops / RevOps":
        return "Resume Test B - Sales Ops RevOps"
    if target_track == "Business Analyst":
        return "Resume Test C - Business Analyst"
    if target_track == "Data / Reporting":
        return "Resume Test D - Analytics"
    if target_track == "Operations / Supply Chain":
        return "Resume Test E - Operations"
    return "Resume Test F - General"

def score_job(job):
    job_text = " ".join([
        job.get("job_title", ""),
        job.get("company", ""),
        job.get("location", "")
    ]).lower()

    full_text = " ".join([
        job_text,
        job.get("email_subject", ""),
        job.get("apply_url", ""),
        job.get("source_email_url", "")
    ]).lower()

    score = 0
    reasons = []

    for keyword, points in TARGET_KEYWORDS.items():
        if keyword in job_text:
            score += points
            reasons.append(keyword)

    # Resume-aligned bonuses based on Domonique's strongest proof points.
    resume_aligned_terms = {
        "flow": 15,
        "flow builder": 20,
        "reports": 12,
        "dashboards": 12,
        "uat": 10,
        "user acceptance": 10,
        "requirements": 12,
        "stakeholder": 10,
        "process improvement": 15,
        "crm data": 10,
        "case management": 10,
        "support queue": 10,
        "salesforce admin": 20,
        "salesforce administrator": 25,
        "agentforce": 15,
        "maximo": 12,
        "inventory": 12,
        "procurement": 10,
    }
    for term, points in resume_aligned_terms.items():
        if term in full_text:
            score += points
            reasons.append(f"Resume aligned: {term}")

    # Apply-link bonus: jobs with direct links are faster to act on.
    if job.get("apply_url"):
        score += 10
        reasons.append("Apply link found")

    location_status = get_location_status(job_text)

    if "remote" in job_text or "work from home" in job_text:
        score += 40
        reasons.append("Remote")

    if "hybrid" in job_text:
        score += 30
        reasons.append("Hybrid")

    if any(city in job_text for city in MN_CITIES) or ", mn" in job_text or "minnesota" in job_text:
        score += 15
        reasons.append("Minnesota location")

    if location_status == "Outside Target Area":
        score -= 75
        reasons.append("Outside target location")

    if location_status == "On-site":
        score -= 30
        reasons.append("On-site role")

    company_bonus, company_match, company_category = get_company_bonus(job.get("company", ""))
    if company_bonus > 0:
        score += company_bonus
        reasons.append(f"{company_category}: {company_match}")

    if has_salary_signal(full_text):
        if re.search(r"\$?\s?(100|110|120|130|140|150|160|170|180|190|200)\s?k", full_text):
            score += 25
            reasons.append("100k+ salary signal")
        elif re.search(r"\$?\s?(80|90)\s?k", full_text):
            score += 15
            reasons.append("80k+ salary signal")

    red_flags = [
        "senior director", "director", "nurse", "retail associate",
        "customer service rep", "warehouse associate", "cashier",
        "restaurant", "driver", "cdl"
    ]

    if any(flag in job_text for flag in red_flags):
        score -= 25
        reasons.append("Possible mismatch")

    if score < 0:
        score = 0

    target_track = get_target_track(job)
    target_strength = get_target_strength(job)

    job["score"] = score
    job["match_percent"] = get_match_percent(score)
    job["fit_category"] = get_fit_category(score)
    job["target_track"] = target_track
    job["recommended_resume"] = get_recommended_resume(target_track)
    job["resume_file_suggestion"] = get_resume_file_suggestion(target_track)
    job["resume_test_group"] = get_resume_test_group(target_track)
    job["location_status"] = location_status
    job["company_bonus"] = company_bonus
    job["company_watchlist_category"] = company_category
    job["target_strength"] = target_strength
    job["recommendation"] = get_recommendation(score)
    job["apply_priority"] = get_apply_priority(score)
    job["match_reasons"] = ", ".join(reasons)

    resume_advice = get_resume_advisor(job)
    job.update(resume_advice)

    job_intelligence = get_job_intelligence(job)
    job.update(job_intelligence)

    job["application_status"] = "Not Applied"
    job["application_date"] = ""
    job["interview_date"] = ""
    job["follow_up_date"] = ""
    job["apply_url"] = job.get("apply_url", "")
    job["job_link"] = job.get("job_link", job.get("apply_url", ""))
    job["source_email_url"] = job.get("source_email_url", "")
    job["resume_file_suggestion"] = job.get("resume_file_suggestion", get_resume_file_suggestion(target_track))
    job["resume_test_group"] = job.get("resume_test_group", get_resume_test_group(target_track))
    job["notes"] = ""

    return job

def create_job_key(job):
    """Create a stable key so JobRadar can remember the same job across runs."""
    title = normalize_title(job.get("job_title", ""))
    company = normalize_company(job.get("company", ""))
    location = normalize_location(job.get("location", ""))
    return f"{title}|{company}|{location}"


def load_application_tracker():
    """Load persistent application tracking data or create a blank tracker."""
    os.makedirs("data", exist_ok=True)

    if not os.path.exists(TRACKER_FILE):
        tracker_df = pd.DataFrame(columns=TRACKER_COLUMNS)
        tracker_df.to_csv(TRACKER_FILE, index=False)
        return tracker_df

    tracker_df = pd.read_csv(TRACKER_FILE, dtype=str).fillna("")

    for column in TRACKER_COLUMNS:
        if column not in tracker_df.columns:
            tracker_df[column] = ""

    tracker_df = tracker_df[TRACKER_COLUMNS]
    tracker_df = tracker_df.drop_duplicates(subset=["job_key"], keep="last")
    return tracker_df


def save_application_tracker(tracker_df):
    """Save the persistent application tracker."""
    os.makedirs("data", exist_ok=True)
    tracker_df = tracker_df[TRACKER_COLUMNS]
    tracker_df.to_csv(TRACKER_FILE, index=False)


def apply_application_tracker(scored_jobs):
    """Merge saved application statuses into the current JobRadar results."""
    tracker_df = load_application_tracker()
    today = datetime.now().strftime("%Y-%m-%d")

    tracker_lookup = {
        row["job_key"]: row.to_dict()
        for _, row in tracker_df.iterrows()
        if row.get("job_key", "")
    }

    new_tracker_rows = []

    for job in scored_jobs:
        job_key = create_job_key(job)
        saved = tracker_lookup.get(job_key)

        if saved:
            status = saved.get("status", "Not Applied") or "Not Applied"
            application_stage = saved.get("application_stage", status) or status
            date_applied = saved.get("date_applied", "")
            follow_up = saved.get("follow_up", "")
            last_updated = saved.get("last_updated", "")
            resume_used = saved.get("resume_used", "")
            cover_letter_used = saved.get("cover_letter_used", "")
            recruiter_name = saved.get("recruiter_name", "")
            recruiter_email = saved.get("recruiter_email", "")
            interview_date = saved.get("interview_date", "")
            salary = saved.get("salary", "")
            saved_job_link = saved.get("job_link", "")
            apply_url = saved.get("apply_url", "") or job.get("apply_url", "") or saved_job_link
            source_email_url = saved.get("source_email_url", "") or job.get("source_email_url", "")
            job_link = apply_url or saved_job_link
            resume_file_suggestion = saved.get("resume_file_suggestion", "") or job.get("resume_file_suggestion", "")
            resume_test_group = saved.get("resume_test_group", "") or job.get("resume_test_group", "")
            notes = saved.get("notes", "")
        else:
            status = "Not Applied"
            application_stage = "Not Applied"
            date_applied = ""
            follow_up = ""
            last_updated = today
            resume_used = job.get("recommended_resume", "")
            cover_letter_used = "No"
            recruiter_name = ""
            recruiter_email = ""
            interview_date = ""
            salary = ""
            apply_url = job.get("apply_url", "")
            source_email_url = job.get("source_email_url", "")
            job_link = apply_url
            resume_file_suggestion = job.get("resume_file_suggestion", "")
            resume_test_group = job.get("resume_test_group", "")
            notes = ""

            new_tracker_rows.append({
                "job_key": job_key,
                "job_title": job.get("job_title", ""),
                "company": job.get("company", ""),
                "status": status,
                "application_stage": application_stage,
                "date_applied": date_applied,
                "follow_up": follow_up,
                "last_updated": last_updated,
                "resume_used": resume_used,
                "cover_letter_used": cover_letter_used,
                "recruiter_name": recruiter_name,
                "recruiter_email": recruiter_email,
                "interview_date": interview_date,
                "salary": salary,
                "job_link": job_link,
                "apply_url": apply_url,
                "source_email_url": source_email_url,
                "resume_file_suggestion": resume_file_suggestion,
                "resume_test_group": resume_test_group,
                "notes": notes,
            })

        job["job_key"] = job_key
        job["application_status"] = status
        job["application_stage"] = application_stage
        job["application_date"] = date_applied
        job["follow_up_date"] = follow_up
        job["last_updated"] = last_updated
        job["resume_used"] = resume_used
        job["cover_letter_used"] = cover_letter_used
        job["recruiter_name"] = recruiter_name
        job["recruiter_email"] = recruiter_email
        job["interview_date"] = interview_date
        job["salary"] = salary
        job["job_link"] = job_link
        job["apply_url"] = apply_url
        job["source_email_url"] = source_email_url
        job["resume_file_suggestion"] = resume_file_suggestion
        job["resume_test_group"] = resume_test_group
        job["notes"] = notes

    if new_tracker_rows:
        tracker_df = pd.concat(
            [tracker_df, pd.DataFrame(new_tracker_rows)],
            ignore_index=True
        )
        tracker_df = tracker_df.drop_duplicates(subset=["job_key"], keep="last")
        save_application_tracker(tracker_df)

    return scored_jobs



def normalize_company(company):
    company = str(company or "").lower().strip()
    company = re.sub(r"\b(inc|llc|ltd|corp|corporation|company|co)\b", "", company)
    company = re.sub(r"[^a-z0-9]+", " ", company)
    return re.sub(r"\s+", " ", company).strip()


def normalize_location(location):
    location = str(location or "").lower().strip()
    location = location.replace("united states", "us")
    location = location.replace("work from home", "remote")
    location = re.sub(r"[^a-z0-9, ]+", " ", location)
    return re.sub(r"\s+", " ", location).strip()

def normalize_title(title):
    title = title.lower()
    title = re.sub(r"\s+", " ", title)
    title = title.replace(" ii", "")
    title = title.replace(" iii", "")
    title = title.replace(" sr.", " senior")
    return title.strip()


def remove_duplicates(jobs):
    seen = set()
    unique_jobs = []

    for job in jobs:
        title = normalize_title(job.get("job_title", ""))
        company = normalize_company(job.get("company", ""))
        location = normalize_location(job.get("location", ""))
        key = (title, company, location)

        if key not in seen:
            seen.add(key)
            unique_jobs.append(job)

    return unique_jobs


def autosize_columns(worksheet):
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 45)


def style_worksheet(worksheet):
    header_fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")

    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    worksheet.freeze_panes = "A2"
    autosize_columns(worksheet)



def apply_excel_hyperlinks(worksheet):
    """Turn URL columns into friendly clickable links in Excel."""
    headers = {cell.value: cell.column for cell in worksheet[1] if cell.value}

    hyperlink_columns = {
        "apply_url": "Apply",
        "job_link": "Apply",
        "source_email_url": "Open Email",
    }

    for column_name, label in hyperlink_columns.items():
        if column_name not in headers:
            continue

        col_idx = headers[column_name]
        for row_idx in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            url = str(cell.value or "").strip()
            if url.startswith("http"):
                cell.hyperlink = url
                cell.value = label
                cell.style = "Hyperlink"


def style_workbook(workbook):
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        style_worksheet(worksheet)
        apply_excel_hyperlinks(worksheet)

def build_dashboard_df(df, recruiter_df):
    metrics = [
        ("Total Jobs", len(df)),
        ("Apply Now", len(df[df["recommendation"] == "Apply Now"])),
        ("High Priority", len(df[df["recommendation"] == "High Priority"])),
        ("Review", len(df[df["recommendation"] == "Review"])),
        ("Skip", len(df[df["recommendation"] == "Skip"])),
        ("Weekend Apply List", len(df[df["recommendation"].isin(["Apply Now", "High Priority"])])),
        ("Apply Links Found", len(df[df.get("apply_url", "").fillna("").astype(str).str.startswith("http")]) if "apply_url" in df.columns else 0),
        ("Source Email Links", len(df[df.get("source_email_url", "").fillna("").astype(str).str.startswith("http")]) if "source_email_url" in df.columns else 0),
        ("Recruiter Leads", len(recruiter_df)),
        ("Preferred Location Jobs", len(df[df["location_status"] == "Preferred"]) if "location_status" in df.columns else 0),
        ("Outside Target Area", len(df[df["location_status"] == "Outside Target Area"]) if "location_status" in df.columns else 0),
        ("Applications Submitted", len(df[df["application_stage"] == "Applied"]) if "application_stage" in df.columns else 0),
        ("Recruiter Contacted", len(df[df["application_stage"] == "Recruiter Contacted"]) if "application_stage" in df.columns else 0),
        ("Phone Screens", len(df[df["application_stage"] == "Phone Screen"]) if "application_stage" in df.columns else 0),
        ("Interviews", len(df[df["application_stage"].isin(["Interview Scheduled", "Interview Complete", "Final Interview"])]) if "application_stage" in df.columns else 0),
        ("Offers", len(df[df["application_stage"] == "Offer"]) if "application_stage" in df.columns else 0),
        ("Accepted", len(df[df["application_stage"] == "Accepted"]) if "application_stage" in df.columns else 0),
        ("Rejected", len(df[df["application_stage"] == "Rejected"]) if "application_stage" in df.columns else 0),
    ]

    return pd.DataFrame(metrics, columns=["Metric", "Count"])



def build_skill_gap_df(df):
    """Summarize recurring missing/watch skills across current job results."""
    skill_counts = {}

    if "missing_or_watch_skills" not in df.columns:
        return pd.DataFrame(columns=["skill", "job_count", "priority", "suggested_action"])

    for skills_text in df["missing_or_watch_skills"].fillna(""):
        skills_text = str(skills_text).strip()
        if not skills_text or skills_text.lower() == "none flagged":
            continue

        for skill in skills_text.split("|"):
            skill = skill.strip()
            if not skill or skill.lower() == "none flagged":
                continue
            skill_counts[skill] = skill_counts.get(skill, 0) + 1

    rows = []
    for skill, count in sorted(skill_counts.items(), key=lambda item: item[1], reverse=True):
        if count >= 8:
            priority = "High"
        elif count >= 4:
            priority = "Medium"
        else:
            priority = "Low"

        rows.append({
            "skill": skill,
            "job_count": count,
            "priority": priority,
            "suggested_action": f"Review or build a small proof-of-skill project for {skill}"
        })

    if not rows:
        rows.append({
            "skill": "No recurring gaps flagged",
            "job_count": 0,
            "priority": "None",
            "suggested_action": "Continue applying and review job descriptions manually for specific requirements"
        })

    return pd.DataFrame(rows)

def export_reports(scored_jobs, recruiter_leads):
    os.makedirs("reports", exist_ok=True)

    df = pd.DataFrame(scored_jobs)
    recruiter_df = pd.DataFrame(recruiter_leads)

    if recruiter_df.empty:
        recruiter_crm_df = load_recruiter_crm()
    else:
        recruiter_df, recruiter_crm_df = apply_recruiter_crm(recruiter_df)

    preferred_columns = [
        "job_key", "job_title", "company", "location", "source",
        "apply_url", "source_email_url", "apply_link_status",
        "score", "match_percent", "fit_category", "target_track",
        "recommended_resume", "resume_fit_level", "resume_fit_score",
        "resume_suggestions", "missing_or_watch_skills", "interview_focus",
        "job_intelligence_summary", "why_this_fits", "resume_focus",
        "job_intelligence_interview_focus", "potential_concerns", "suggested_next_action",
        "location_status", "company_bonus",
        "company_watchlist_category", "target_strength", "recommendation", "apply_priority",
        "application_status", "application_stage", "application_date", "resume_used",
        "resume_file_suggestion", "resume_test_group",
        "cover_letter_used", "recruiter_name", "recruiter_email", "interview_date",
        "salary", "follow_up_date", "last_updated", "job_link", "notes",
        "match_reasons", "email_subject", "email_date"
    ]

    recruiter_columns = [
        "recruiter_key", "recruiter_name", "company", "email",
        "lead_type", "sender", "subject", "date",
        "priority", "status", "crm_status", "relationship_score",
        "suggested_action", "snippet"
    ]

    df = df[[col for col in preferred_columns if col in df.columns]]

    if recruiter_df.empty:
        recruiter_df = pd.DataFrame(columns=recruiter_columns)
    else:
        recruiter_df = recruiter_df[[col for col in recruiter_columns if col in recruiter_df.columns]]

    all_path = "reports/job_radar_results.csv"
    apply_path = "reports/apply_now.csv"
    apply_excel_path = "reports/apply_now.xlsx"
    high_path = "reports/high_priority.csv"
    review_path = "reports/review.csv"
    skip_path = "reports/skip.csv"
    weekend_path = "reports/weekend_apply_list.csv"
    recruiter_path = "reports/recruiter_leads.csv"
    recruiter_crm_path = "reports/recruiter_crm.csv"
    recruiter_crm_excel_path = "reports/recruiter_crm.xlsx"
    tracker_path = "reports/application_tracker_export.csv"
    settings_path = "reports/settings_snapshot.csv"
    company_watchlist_path = "reports/company_watchlist.csv"
    job_intelligence_path = "reports/job_intelligence.csv"
    job_intelligence_excel_path = "reports/job_intelligence.xlsx"
    skill_gap_path = "reports/skill_gap_analysis.csv"
    skill_gap_excel_path = "reports/skill_gap_analysis.xlsx"
    excel_path = "reports/job_radar.xlsx"

    apply_df = df[df["recommendation"] == "Apply Now"]
    high_df = df[df["recommendation"] == "High Priority"]
    review_df = df[df["recommendation"] == "Review"]
    skip_df = df[df["recommendation"] == "Skip"]
    weekend_df = df[df["recommendation"].isin(["Apply Now", "High Priority"])]
    dashboard_df = build_dashboard_df(df, recruiter_df)
    job_intelligence_columns = [
        "job_title", "company", "location", "source", "apply_url", "source_email_url",
        "score", "fit_category", "target_track", "recommended_resume",
        "resume_fit_level", "job_intelligence_summary", "why_this_fits",
        "resume_focus", "job_intelligence_interview_focus",
        "potential_concerns", "suggested_next_action",
        "application_stage", "email_date"
    ]
    job_intelligence_df = df[[col for col in job_intelligence_columns if col in df.columns]]
    skill_gap_df = build_skill_gap_df(df)

    df.to_csv(all_path, index=False)
    apply_df.to_csv(apply_path, index=False)
    high_df.to_csv(high_path, index=False)
    review_df.to_csv(review_path, index=False)
    skip_df.to_csv(skip_path, index=False)
    weekend_df.to_csv(weekend_path, index=False)
    job_intelligence_df.to_csv(job_intelligence_path, index=False)
    skill_gap_df.to_csv(skill_gap_path, index=False)
    recruiter_df.to_csv(recruiter_path, index=False)
    recruiter_crm_df.to_csv(recruiter_crm_path, index=False)
    tracker_df = load_application_tracker()
    tracker_df.to_csv(tracker_path, index=False)
    settings = load_settings()
    settings_df = pd.DataFrame([{"setting": key, "value": json.dumps(value) if isinstance(value, list) else value} for key, value in settings.items()])
    settings_df.to_csv(settings_path, index=False)
    company_watchlist_df = load_company_watchlist()
    company_watchlist_df.to_csv(company_watchlist_path, index=False)

    with pd.ExcelWriter(apply_excel_path, engine="openpyxl") as writer:
        apply_df.to_excel(writer, sheet_name="Apply Now", index=False)
        workbook = writer.book
        style_workbook(workbook)

    with pd.ExcelWriter(job_intelligence_excel_path, engine="openpyxl") as writer:
        job_intelligence_df.to_excel(writer, sheet_name="Job Intelligence", index=False)
        workbook = writer.book
        style_workbook(workbook)

    with pd.ExcelWriter(skill_gap_excel_path, engine="openpyxl") as writer:
        skill_gap_df.to_excel(writer, sheet_name="Skill Gap Analysis", index=False)
        workbook = writer.book
        style_workbook(workbook)

    with pd.ExcelWriter(recruiter_crm_excel_path, engine="openpyxl") as writer:
        recruiter_crm_df.to_excel(writer, sheet_name="Recruiter CRM", index=False)
        workbook = writer.book
        style_workbook(workbook)

    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        dashboard_df.to_excel(writer, sheet_name="Dashboard", index=False)
        weekend_df.to_excel(writer, sheet_name="Weekend Apply List", index=False)
        apply_df.to_excel(writer, sheet_name="Apply Now", index=False)
        high_df.to_excel(writer, sheet_name="High Priority", index=False)
        review_df.to_excel(writer, sheet_name="Review", index=False)
        skip_df.to_excel(writer, sheet_name="Skip", index=False)
        recruiter_df.to_excel(writer, sheet_name="Recruiter Leads", index=False)
        recruiter_crm_df.to_excel(writer, sheet_name="Recruiter CRM", index=False)
        tracker_df.to_excel(writer, sheet_name="Application Tracker", index=False)
        company_watchlist_df.to_excel(writer, sheet_name="Company Watchlist", index=False)
        settings_df.to_excel(writer, sheet_name="Settings", index=False)
        job_intelligence_df.to_excel(writer, sheet_name="Job Intelligence", index=False)
        skill_gap_df.to_excel(writer, sheet_name="Skill Gap Analysis", index=False)
        df.to_excel(writer, sheet_name="All Jobs", index=False)

        workbook = writer.book
        style_workbook(workbook)

    return {
        "all": all_path,
        "apply": apply_path,
        "apply_excel": apply_excel_path,
        "high": high_path,
        "review": review_path,
        "skip": skip_path,
        "weekend": weekend_path,
        "recruiter": recruiter_path,
        "recruiter_crm": recruiter_crm_path,
        "recruiter_crm_excel": recruiter_crm_excel_path,
        "tracker": tracker_path,
        "settings": settings_path,
        "company_watchlist": company_watchlist_path,
        "job_intelligence": job_intelligence_path,
        "job_intelligence_excel": job_intelligence_excel_path,
        "skill_gap": skill_gap_path,
        "skill_gap_excel": skill_gap_excel_path,
        "excel": excel_path,
        "total": len(df),
        "apply_count": len(apply_df),
        "high_count": len(high_df),
        "review_count": len(review_df),
        "skip_count": len(skip_df),
        "weekend_count": len(weekend_df),
        "recruiter_count": len(recruiter_df),
    }


def main():
    print("\nConnecting to Gmail...\n")
    service = connect_gmail()
    print("✅ Gmail Connected Successfully!\n")

    messages = search_job_alert_emails(service)
    print(f"Found {len(messages)} job alert emails from the last 14 days.\n")

    all_jobs = []

    for index, msg in enumerate(messages, start=1):
        email = read_full_email(service, msg["id"])
        print(f"Reading job email #{index}: {email['subject']}")

        sender = email["sender"].lower()

        if "linkedin" in sender:
            extracted_jobs = extract_linkedin_jobs(email)
        elif "google" in sender:
            extracted_jobs = extract_google_jobs(email)
        else:
            extracted_jobs = extract_simple_jobs(email)

        all_jobs.extend(extracted_jobs)

    recruiter_messages = search_recruiter_emails(service)
    print(f"\nFound {len(recruiter_messages)} possible recruiter/networking emails from the last 14 days.\n")

    recruiter_leads = []

    for index, msg in enumerate(recruiter_messages, start=1):
        email = read_full_email(service, msg["id"])
        lead = extract_recruiter_lead(email)

        if lead:
            print(f"Reading recruiter lead #{index}: {email['subject']}")
            recruiter_leads.append(lead)

    all_jobs = remove_duplicates(all_jobs)
    scored_jobs = [score_job(job) for job in all_jobs]
    scored_jobs = apply_application_tracker(scored_jobs)
    scored_jobs = sorted(scored_jobs, key=lambda x: x["score"], reverse=True)

    export_info = export_reports(scored_jobs, recruiter_leads)

    print("\n✅ Job extraction complete.")
    print(f"Found {export_info['total']} unique jobs.")
    print(f"Apply Now: {export_info['apply_count']}")
    print(f"High Priority: {export_info['high_count']}")
    print(f"Review: {export_info['review_count']}")
    print(f"Skip: {export_info['skip_count']}")
    print(f"Weekend Apply List: {export_info['weekend_count']}")
    print(f"Recruiter Leads: {export_info['recruiter_count']}\n")

    print("Reports created:")
    print(f"- {export_info['all']}")
    print(f"- {export_info['apply']}")
    print(f"- {export_info['apply_excel']}")
    print(f"- {export_info['high']}")
    print(f"- {export_info['review']}")
    print(f"- {export_info['skip']}")
    print(f"- {export_info['weekend']}")
    print(f"- {export_info['recruiter']}")
    print(f"- {export_info['recruiter_crm']}")
    print(f"- {export_info['recruiter_crm_excel']}")
    print(f"- {export_info['tracker']}")
    print(f"- {export_info['settings']}")
    print(f"- {export_info['company_watchlist']}")
    print(f"- {export_info['job_intelligence']}")
    print(f"- {export_info['job_intelligence_excel']}")
    print(f"- {export_info['skill_gap']}")
    print(f"- {export_info['skill_gap_excel']}")
    print(f"- {export_info['excel']}\n")

    print("Top Matches:\n")

    for job in scored_jobs[:15]:
        print("-" * 80)
        print(f"Title: {job['job_title']}")
        print(f"Company: {job['company']}")
        print(f"Location: {job['location']}")
        print(f"Source: {job['source']}")
        print(f"Apply URL: {job.get('apply_url', '') or job.get('source_email_url', '')}")
        print(f"Score: {job['score']}")
        print(f"Match %: {job['match_percent']}%")
        print(f"Fit Category: {job['fit_category']}")
        print(f"Target Track: {job['target_track']}")
        print(f"Recommended Resume: {job['recommended_resume']}")
        print(f"Resume Fit: {job.get('resume_fit_level', '')}")
        print(f"Recommendation: {job['recommendation']}")
        print(f"Apply Priority: {job['apply_priority']}")
        print(f"Application Stage: {job.get('application_stage', '')}")
        print(f"Resume Used: {job.get('resume_used', job.get('recommended_resume', ''))}")
        print(f"Reasons: {job['match_reasons']}")
        print("-" * 80)


if __name__ == "__main__":
    main()
