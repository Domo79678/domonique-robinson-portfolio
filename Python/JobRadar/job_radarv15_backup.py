import os
import re
import base64
from email.utils import parsedate_to_datetime

import pandas as pd
from bs4 import BeautifulSoup

from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build

from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

SCOPES = ["https://www.googleapis.com/auth/gmail.readonly"]

BLACKLIST = [
    "add widget", "manage alerts", "manage job alerts", "view all jobs",
    "unsubscribe", "this email was intended", "linkedin premium",
    "you have an invitation", "thanks for being a valued member",
    "see more jobs", "recommended jobs", "job alert", "your job alert",
    "actively recruiting", "view job", "privacy policy", "email preferences",
]

TARGET_KEYWORDS = {
    "salesforce administrator": 45, "salesforce admin": 45, "salesforce": 35,
    "crm administrator": 35, "crm admin": 35, "crm": 25,
    "business analyst": 35, "analyst": 20, "data analyst": 30,
    "reporting": 25, "dashboard": 20, "analytics": 25,
    "operations analyst": 30, "operations": 20,
    "sales operations": 35, "sales ops": 35,
    "revops": 35, "revenue operations": 35,
    "tableau": 20, "sql": 20, "soql": 20, "excel": 15,
    "d365": 15, "dynamics": 15, "erp": 10,
    "supply chain": 20, "inventory": 20, "materials": 15,
    "procurement": 15, "purchasing": 15, "planner": 15,
    "process improvement": 15, "workflow": 10,
}

MN_CITIES = [
    "minneapolis", "st paul", "saint paul", "bloomington", "hopkins",
    "shoreview", "brooklyn park", "arden hills", "plymouth", "minnetonka",
    "roseville", "eden prairie", "maple grove", "st. louis park", "wayzata",
]

PREFERRED_LOCATION_TERMS = [
    "remote", "hybrid", "united states", "usa", "u.s.", "minneapolis",
    "st paul", "saint paul", "bloomington", "hopkins", "shoreview",
    "brooklyn park", "arden hills", "plymouth", "minnetonka", "roseville",
    "eden prairie", "maple grove", "st. louis park", "wayzata", "mn"
]

NON_US_LOCATION_TERMS = [
    "india", "france", "australia", "japan", "south korea", "korea",
    "indonesia", "mexico", "netherlands", "united kingdom", "nigeria",
    "guam", "canada", "germany", "brazil", "philippines", "singapore"
]

TARGET_COMPANIES = {
    "optum": 25,
    "unitedhealth": 25,
    "unitedhealth group": 25,
    "cargill": 20,
    "best buy": 20,
    "thermo fisher": 20,
    "salesforce": 25,
    "medica": 15,
    "wells fargo": 15,
    "patterson companies": 15,
    "robert half": 10,
    "on-demand group": 10,
    "srs acquiom": 10,
    "catch resource management": 10,
}


def connect_gmail():
    creds = None

    if os.path.exists("token.json"):
        creds = Credentials.from_authorized_user_file("token.json", SCOPES)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file("credentials.json", SCOPES)
            creds = flow.run_local_server(port=0)

        with open("token.json", "w") as token:
            token.write(creds.to_json())

    return build("gmail", "v1", credentials=creds)


def get_header(headers, name):
    for header in headers:
        if header["name"].lower() == name.lower():
            return header["value"]
    return ""


def search_job_alert_emails(service):
    query = (
        "("
        "from:jobalerts-noreply@linkedin.com OR "
        "from:jobalerts@sites.careerbuilder.com OR "
        "from:notify-noreply@google.com OR "
        "from:googlealerts-noreply@google.com OR "
        "from:alerts@indeed.com OR "
        "from:noreply@indeed.com OR "
        "from:ziprecruiter.com"
        ") newer_than:14d"
    )

    results = service.users().messages().list(
        userId="me",
        q=query,
        maxResults=100
    ).execute()

    return results.get("messages", [])


def search_recruiter_emails(service):
    query = (
        "("
        "from:messages-noreply@linkedin.com OR "
        "from:messaging-digest-noreply@linkedin.com OR "
        "from:invitations@linkedin.com OR "
        "from:linkedin.com"
        ") "
        "("
        "recruiter OR hiring OR opportunity OR interview OR "
        '"messaged you" OR "sent you an invitation" OR "waiting for your response"'
        ") newer_than:14d"
    )

    results = service.users().messages().list(
        userId="me",
        q=query,
        maxResults=50
    ).execute()

    return results.get("messages", [])


def decode_body(data):
    if not data:
        return ""

    data = data.replace("-", "+").replace("_", "/")
    decoded_bytes = base64.b64decode(data)
    return decoded_bytes.decode("utf-8", errors="ignore")


def get_email_body(payload):
    body_text = ""

    if "parts" in payload:
        for part in payload["parts"]:
            mime_type = part.get("mimeType", "")
            body_data = part.get("body", {}).get("data")

            if mime_type in ["text/html", "text/plain"] and body_data:
                body_text += decode_body(body_data)

            if "parts" in part:
                body_text += get_email_body(part)
    else:
        body_data = payload.get("body", {}).get("data")
        if body_data:
            body_text += decode_body(body_data)

    return body_text


def read_full_email(service, message_id):
    message = service.users().messages().get(
        userId="me",
        id=message_id,
        format="full"
    ).execute()

    headers = message["payload"]["headers"]

    subject = get_header(headers, "Subject")
    sender = get_header(headers, "From")
    date_raw = get_header(headers, "Date")

    try:
        email_date = parsedate_to_datetime(date_raw).strftime("%Y-%m-%d")
    except Exception:
        email_date = date_raw

    raw_body = get_email_body(message["payload"])
    soup = BeautifulSoup(raw_body, "html.parser")
    clean_text = soup.get_text("\n")

    return {
        "subject": subject,
        "sender": sender,
        "date": email_date,
        "body": clean_text,
        "snippet": message.get("snippet", "")
    }


def is_junk_title(title):
    title_lower = title.lower().strip()

    if len(title_lower) < 4:
        return True

    if set(title_lower) <= {"-", "_", "—", " "}:
        return True

    if title_lower.count("-") > 10:
        return True

    return any(bad_word in title_lower for bad_word in BLACKLIST)


def build_job(job_title, company, location, source, email):
    return {
        "job_title": job_title.strip(),
        "company": company.strip(),
        "location": location.strip(),
        "source": source,
        "email_subject": email["subject"],
        "email_date": email["date"],
        "application_status": "Not Applied",
        "application_date": "",
        "interview_date": "",
        "follow_up_date": "",
        "job_link": "",
        "notes": ""
    }


def extract_linkedin_jobs(email):
    jobs = []
    lines = [line.strip() for line in email["body"].splitlines() if line.strip()]

    for i, line in enumerate(lines):
        if is_junk_title(line):
            continue

        next_line = lines[i + 1] if i + 1 < len(lines) else ""

        if "·" in next_line:
            parts = [p.strip() for p in next_line.split("·")]
            company = parts[0] if len(parts) > 0 else ""
            location = parts[1] if len(parts) > 1 else ""

            if not is_junk_title(company):
                jobs.append(build_job(line, company, location, "LinkedIn", email))

    return jobs


def extract_google_jobs(email):
    jobs = []
    lines = [line.strip() for line in email["body"].splitlines() if line.strip()]

    for i, line in enumerate(lines):
        if is_junk_title(line):
            continue

        if any(keyword in line.lower() for keyword in [
            "analyst", "administrator", "salesforce", "crm", "operations",
            "data", "reporting", "manager", "coordinator", "specialist",
            "inventory", "supply chain"
        ]):
            company = lines[i + 1] if i + 1 < len(lines) else ""
            location = lines[i + 2] if i + 2 < len(lines) else ""

            if not is_junk_title(company):
                jobs.append(build_job(line, company, location, "Google Job Alert", email))

    return jobs


def extract_simple_jobs(email):
    jobs = []
    lines = [line.strip() for line in email["body"].splitlines() if line.strip()]

    for line in lines:
        if is_junk_title(line):
            continue

        if any(keyword in line.lower() for keyword in TARGET_KEYWORDS.keys()):
            jobs.append(build_job(line, "", "", "Other", email))

    return jobs


def extract_recruiter_lead(email):
    subject = email.get("subject", "")
    sender = email.get("sender", "")
    snippet = email.get("snippet", "")
    body = email.get("body", "")

    combined = " ".join([subject, sender, snippet, body]).lower()

    if "job alert" in combined or "recommended jobs" in combined:
        return None

    if not any(word in combined for word in [
        "recruiter", "hiring", "opportunity", "interview",
        "messaged you", "sent you an invitation", "waiting for your response"
    ]):
        return None

    lead_type = "LinkedIn Message"
    if "invitation" in combined or "waiting for your response" in combined:
        lead_type = "LinkedIn Invitation"

    priority = "Review"
    if any(word in combined for word in ["recruiter", "hiring", "opportunity", "interview"]):
        priority = "High Priority"

    return {
        "lead_type": lead_type,
        "sender": sender,
        "subject": subject,
        "date": email.get("date", ""),
        "priority": priority,
        "status": "Not Reviewed",
        "snippet": snippet
    }


def has_salary_signal(text):
    return bool(
        re.search(r"\$\s?(80|90|100|110|120|130|140|150|160|170|180|190|200)\s?k", text)
        or re.search(r"\$\s?(80,000|90,000|100,000|110,000|120,000|130,000|140,000|150,000)", text)
        or re.search(r"up to \$\s?\d+", text)
    )


def get_target_track(job):
    text = " ".join([
        job.get("job_title", ""),
        job.get("company", ""),
        job.get("location", ""),
        job.get("email_subject", "")
    ]).lower()

    if any(word in text for word in ["salesforce", "crm"]):
        return "Salesforce / CRM"

    if any(word in text for word in ["sales operations", "sales ops", "revops", "revenue operations"]):
        return "Sales Ops / RevOps"

    if any(word in text for word in ["business analyst", "project coordinator", "process analyst", "product owner"]):
        return "Business Analyst"

    if any(word in text for word in ["data analyst", "analytics", "reporting", "dashboard", "tableau", "sql", "excel"]):
        return "Data / Reporting"

    if any(word in text for word in [
        "operations", "supply chain", "inventory", "materials",
        "procurement", "purchasing", "planner", "logistics"
    ]):
        return "Operations / Supply Chain"

    return "General / Other"


def get_recommended_resume(target_track):
    if target_track == "Salesforce / CRM":
        return "Salesforce Resume"
    if target_track == "Sales Ops / RevOps":
        return "Sales Ops / RevOps Resume"
    if target_track == "Business Analyst":
        return "Business Analyst Resume"
    if target_track == "Data / Reporting":
        return "Business Analytics Resume"
    if target_track == "Operations / Supply Chain":
        return "Operations / Supply Chain Resume"
    return "General Resume"



def get_location_status(job):
    text = " ".join([
        job.get("job_title", ""),
        job.get("company", ""),
        job.get("location", ""),
        job.get("email_subject", "")
    ]).lower()

    if any(term in text for term in PREFERRED_LOCATION_TERMS):
        return "Preferred"

    if any(term in text for term in NON_US_LOCATION_TERMS):
        return "Outside Target Area"

    return "Unknown"


def get_company_score(job):
    text = " ".join([
        job.get("company", ""),
        job.get("job_title", ""),
        job.get("email_subject", "")
    ]).lower()

    for company, points in TARGET_COMPANIES.items():
        if company in text:
            return points

    return 0


def get_company_tier(company_score):
    if company_score >= 20:
        return "Strong Target Company"
    if company_score >= 10:
        return "Good Target Company"
    return "Standard Company"

def get_match_percent(score):
    return min(100, score)


def get_fit_category(score):
    if score >= 90:
        return "Excellent Match"
    if score >= 70:
        return "Strong Match"
    if score >= 50:
        return "Possible Match"
    return "Low Match"


def get_recommendation(score):
    if score >= 90:
        return "Apply Now"
    if score >= 70:
        return "High Priority"
    if score >= 50:
        return "Review"
    return "Skip"


def get_apply_priority(score):
    if score >= 90:
        return "Apply within 24 hours"
    if score >= 70:
        return "Apply this week"
    if score >= 50:
        return "Review if time"
    return "Do not prioritize"


def score_job(job):
    job_text = " ".join([
        job.get("job_title", ""),
        job.get("company", ""),
        job.get("location", "")
    ]).lower()

    full_text = " ".join([
        job_text,
        job.get("email_subject", "")
    ]).lower()

    score = 0
    reasons = []

    for keyword, points in TARGET_KEYWORDS.items():
        if keyword in job_text:
            score += points
            reasons.append(keyword)

    if "remote" in job_text:
        score += 40
        reasons.append("Remote")

    if "hybrid" in job_text:
        score += 30
        reasons.append("Hybrid")

    if any(city in job_text for city in MN_CITIES) or ", mn" in job_text:
        score += 15
        reasons.append("Minnesota location")

    location_status = get_location_status(job)

    if location_status == "Preferred":
        score += 10
        reasons.append("Preferred location")

    if location_status == "Outside Target Area":
        score -= 35
        reasons.append("Outside target area")

    company_score = get_company_score(job)

    if company_score > 0:
        score += company_score
        reasons.append("Target company")

    if has_salary_signal(full_text):
        if re.search(r"\$?\s?(100|110|120|130|140|150|160|170|180|190|200)\s?k", full_text):
            score += 25
            reasons.append("100k+ salary signal")
        elif re.search(r"\$?\s?(80|90)\s?k", full_text):
            score += 15
            reasons.append("80k+ salary signal")

    red_flags = [
        "senior director", "director", "nurse", "retail associate",
        "customer service rep", "warehouse associate", "cashier",
        "restaurant", "driver", "cdl"
    ]

    if any(flag in job_text for flag in red_flags):
        score -= 25
        reasons.append("Possible mismatch")

    if score < 0:
        score = 0

    target_track = get_target_track(job)

    job["score"] = score
    job["match_percent"] = get_match_percent(score)
    job["fit_category"] = get_fit_category(score)
    job["target_track"] = target_track
    job["recommended_resume"] = get_recommended_resume(target_track)
    job["location_status"] = get_location_status(job)
    job["company_score"] = get_company_score(job)
    job["company_tier"] = get_company_tier(job["company_score"])
    job["recommendation"] = get_recommendation(score)
    job["apply_priority"] = get_apply_priority(score)
    job["match_reasons"] = ", ".join(reasons)

    return job


def normalize_title(title):
    title = title.lower()
    title = re.sub(r"\s+", " ", title)
    title = title.replace(" ii", "")
    title = title.replace(" iii", "")
    title = title.replace(" sr.", " senior")
    return title.strip()


def remove_duplicates(jobs):
    seen = set()
    unique_jobs = []

    for job in jobs:
        title = normalize_title(job.get("job_title", ""))
        company = job.get("company", "").lower().strip()
        key = (title, company)

        if key not in seen:
            seen.add(key)
            unique_jobs.append(job)

    return unique_jobs


def autosize_columns(worksheet):
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 45)


def style_worksheet(worksheet):
    header_fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")

    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    worksheet.freeze_panes = "A2"
    autosize_columns(worksheet)


def build_dashboard_df(df, recruiter_df):
    metrics = [
        ("Total Jobs", len(df)),
        ("Apply Now", len(df[df["recommendation"] == "Apply Now"])),
        ("High Priority", len(df[df["recommendation"] == "High Priority"])),
        ("Review", len(df[df["recommendation"] == "Review"])),
        ("Skip", len(df[df["recommendation"] == "Skip"])),
        ("Weekend Apply List", len(df[df["recommendation"].isin(["Apply Now", "High Priority"])])),
        ("Recruiter Leads", len(recruiter_df)),
    ]

    return pd.DataFrame(metrics, columns=["Metric", "Count"])


def export_reports(scored_jobs, recruiter_leads):
    os.makedirs("reports", exist_ok=True)

    df = pd.DataFrame(scored_jobs)
    recruiter_df = pd.DataFrame(recruiter_leads)

    preferred_columns = [
        "job_title", "company", "location", "source",
        "score", "match_percent", "fit_category", "target_track",
        "recommended_resume", "location_status", "company_score", "company_tier",
        "recommendation", "apply_priority", "application_status",
        "application_date", "interview_date", "follow_up_date",
        "job_link", "notes", "match_reasons", "email_subject", "email_date"
    ]

    recruiter_columns = [
        "lead_type", "sender", "subject", "date",
        "priority", "status", "snippet"
    ]

    df = df[[col for col in preferred_columns if col in df.columns]]

    if recruiter_df.empty:
        recruiter_df = pd.DataFrame(columns=recruiter_columns)
    else:
        recruiter_df = recruiter_df[[col for col in recruiter_columns if col in recruiter_df.columns]]

    all_path = "reports/job_radar_results.csv"
    apply_path = "reports/apply_now.csv"
    high_path = "reports/high_priority.csv"
    review_path = "reports/review.csv"
    skip_path = "reports/skip.csv"
    weekend_path = "reports/weekend_apply_list.csv"
    recruiter_path = "reports/recruiter_leads.csv"
    excel_path = "reports/job_radar.xlsx"

    apply_df = df[df["recommendation"] == "Apply Now"]
    high_df = df[df["recommendation"] == "High Priority"]
    review_df = df[df["recommendation"] == "Review"]
    skip_df = df[df["recommendation"] == "Skip"]
    weekend_df = df[df["recommendation"].isin(["Apply Now", "High Priority"])]
    dashboard_df = build_dashboard_df(df, recruiter_df)

    df.to_csv(all_path, index=False)
    apply_df.to_csv(apply_path, index=False)
    high_df.to_csv(high_path, index=False)
    review_df.to_csv(review_path, index=False)
    skip_df.to_csv(skip_path, index=False)
    weekend_df.to_csv(weekend_path, index=False)
    recruiter_df.to_csv(recruiter_path, index=False)

    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        dashboard_df.to_excel(writer, sheet_name="Dashboard", index=False)
        weekend_df.to_excel(writer, sheet_name="Weekend Apply List", index=False)
        apply_df.to_excel(writer, sheet_name="Apply Now", index=False)
        high_df.to_excel(writer, sheet_name="High Priority", index=False)
        review_df.to_excel(writer, sheet_name="Review", index=False)
        skip_df.to_excel(writer, sheet_name="Skip", index=False)
        recruiter_df.to_excel(writer, sheet_name="Recruiter Leads", index=False)
        df.to_excel(writer, sheet_name="All Jobs", index=False)

        workbook = writer.book
        for sheet_name in workbook.sheetnames:
            style_worksheet(workbook[sheet_name])

    return {
        "all": all_path,
        "apply": apply_path,
        "high": high_path,
        "review": review_path,
        "skip": skip_path,
        "weekend": weekend_path,
        "recruiter": recruiter_path,
        "excel": excel_path,
        "total": len(df),
        "apply_count": len(apply_df),
        "high_count": len(high_df),
        "review_count": len(review_df),
        "skip_count": len(skip_df),
        "weekend_count": len(weekend_df),
        "recruiter_count": len(recruiter_df),
    }


def main():
    print("\nConnecting to Gmail...\n")
    service = connect_gmail()
    print("✅ Gmail Connected Successfully!\n")

    messages = search_job_alert_emails(service)
    print(f"Found {len(messages)} job alert emails from the last 14 days.\n")

    all_jobs = []

    for index, msg in enumerate(messages, start=1):
        email = read_full_email(service, msg["id"])
        print(f"Reading job email #{index}: {email['subject']}")

        sender = email["sender"].lower()

        if "linkedin" in sender:
            extracted_jobs = extract_linkedin_jobs(email)
        elif "google" in sender:
            extracted_jobs = extract_google_jobs(email)
        else:
            extracted_jobs = extract_simple_jobs(email)

        all_jobs.extend(extracted_jobs)

    recruiter_messages = search_recruiter_emails(service)
    print(f"\nFound {len(recruiter_messages)} possible recruiter/networking emails from the last 14 days.\n")

    recruiter_leads = []

    for index, msg in enumerate(recruiter_messages, start=1):
        email = read_full_email(service, msg["id"])
        lead = extract_recruiter_lead(email)

        if lead:
            print(f"Reading recruiter lead #{index}: {email['subject']}")
            recruiter_leads.append(lead)

    all_jobs = remove_duplicates(all_jobs)
    scored_jobs = [score_job(job) for job in all_jobs]
    scored_jobs = sorted(scored_jobs, key=lambda x: x["score"], reverse=True)

    export_info = export_reports(scored_jobs, recruiter_leads)

    print("\n✅ Job extraction complete.")
    print(f"Found {export_info['total']} unique jobs.")
    print(f"Apply Now: {export_info['apply_count']}")
    print(f"High Priority: {export_info['high_count']}")
    print(f"Review: {export_info['review_count']}")
    print(f"Skip: {export_info['skip_count']}")
    print(f"Weekend Apply List: {export_info['weekend_count']}")
    print(f"Recruiter Leads: {export_info['recruiter_count']}\n")

    print("Reports created:")
    print(f"- {export_info['all']}")
    print(f"- {export_info['apply']}")
    print(f"- {export_info['high']}")
    print(f"- {export_info['review']}")
    print(f"- {export_info['skip']}")
    print(f"- {export_info['weekend']}")
    print(f"- {export_info['recruiter']}")
    print(f"- {export_info['excel']}\n")

    print("Top Matches:\n")

    for job in scored_jobs[:15]:
        print("-" * 80)
        print(f"Title: {job['job_title']}")
        print(f"Company: {job['company']}")
        print(f"Location: {job['location']}")
        print(f"Source: {job['source']}")
        print(f"Score: {job['score']}")
        print(f"Match %: {job['match_percent']}%")
        print(f"Fit Category: {job['fit_category']}")
        print(f"Target Track: {job['target_track']}")
        print(f"Recommended Resume: {job['recommended_resume']}")
        print(f"Recommendation: {job['recommendation']}")
        print(f"Apply Priority: {job['apply_priority']}")
        print(f"Reasons: {job['match_reasons']}")
        print("-" * 80)


if __name__ == "__main__":
    main()
